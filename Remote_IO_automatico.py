import pandas as pd
import xml.etree.ElementTree as ET
import os
import re
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from datetime import datetime
from openpyxl.worksheet.pagebreak import Break # Import necessário para quebras de página
from typing import Dict, Any, List
import sys

#definições ------------------------------------------
ARQUIVO_UNITPRO = 'unitpro.xef'
TIPOS_PERMITIDOS = ['WORD', 'BOOL', 'EBOOL', 'INT']



MODELOS_CAPACIDADE = {
    "140ACI03000": 8, "140ACO02000": 4, "140ACO13000": 8,
    "140ARI03010": 8, "140DDI84100": 32, "140DAI54000": 16,
    "140DAI55300": 32, "140DAO84210": 16, "140DAI74000": 16,
    "140DDI35300": 32, "140DDO35300": 32, "BMXDDI3202K": 32,
    "BMXDDO3202K": 32
}

class Canal:
    def __init__(self, numero):
        self.numero = numero
        self.nome = ""
        self.comentario = ""

class Slot:
    def __init__(self, numero, modelo):
        self.numero = numero
        self.modelo = modelo
        self.qtd_canais = MODELOS_CAPACIDADE.get(modelo, 0)
        # Inicializa a lista de canais com base na capacidade do modelo
        self.canais = [Canal(i+1) for i in range(self.qtd_canais)]

class Drop:
    def __init__(self, numero):
        self.numero = numero
        self.slots = {} # Dicionário {numero_slot: Objeto Slot}

def gerar_matriz_plc(caminho):
    tree = ET.parse(caminho)
    root = tree.getroot()
    
    drops = {} # Dicionário {numero_drop: Objeto Drop}

    # Procurar por módulos (Quantum ou outros que sigam a mesma estrutura)
    # O XPath pode variar dependendo da estrutura completa do seu arquivo
    for module in root.findall(".//moduleQuantum"):
        try:
            # Pegar o Part Number (Modelo)
            part_item = module.find("partItem")
            modelo = part_item.get("partNumber")

            # Pegar o TopoAddress (Ex: \2.2\1.3)
            equip_info = module.find("equipInfo")
            address = equip_info.get("topoAddress")

            # Regex para extrair DROP e SLOT do endereço \2.X\1.Y
            match = re.search(r'\\2\.(\d+)\\1\.(\d+)', address)
            if match:
                num_drop = int(match.group(1))
                num_slot = int(match.group(2))

                # Adiciona Drop se não existir
                if num_drop not in drops:
                    drops[num_drop] = Drop(num_drop)
                
                # Adiciona Slot ao Drop
                drops[num_drop].slots[num_slot] = Slot(num_slot, modelo)
                
                print(f"Mapeado: Drop {num_drop}, Slot {num_slot} -> Modelo {modelo}")

        except Exception as e:
            print(f"Erro ao processar módulo: {e}")

    return drops

def ler_variaveis_unitpro(caminho_arquivo: str) -> List[Dict[str, str]]:
    """Lê todas as variáveis do unitpro.xef."""
    # ... (Implementação omitida por brevidade, assumida como funcional)
    lista_variaveis = []
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()
    except (FileNotFoundError, ET.ParseError) as e:
        print(f"ERRO: Não foi possível ler ou fazer o parse do arquivo: {e}")
        return lista_variaveis

    for var_element in root.findall('.//variables'):
        nome = var_element.get('name')
        tipo = var_element.get('typeName')
        endereco = var_element.get('topologicalAddress')
        comentario_element = var_element.find('comment')
        comentario = comentario_element.text.strip() if comentario_element is not None and comentario_element.text else ""

    

        if nome and tipo in TIPOS_PERMITIDOS: # and endereco
            lista_variaveis.append({
                'nome': nome,
                'comentario': comentario,
                'endereco': endereco,
                'tipo': tipo
            })
    return lista_variaveis

def preencher_canais_da_matriz(caminho_arquivo, matriz_hardware):
    tree = ET.parse(caminho_arquivo)
    root = tree.getroot()

    # Itera sobre todas as variáveis no XML
    for var in root.findall(".//variables"):
        nome_var = var.get("name", "")
        
        # Regex para extrair Drop e Slot do nome (Ex: ED_DROP02_SLOT04)
        match_info = re.search(r'DROP(\d+)_SLOT(\d+)', nome_var)
        
        if match_info:
            num_drop = int(match_info.group(1))
            num_slot = int(match_info.group(2))
            
            # Verifica se esse Drop e Slot existem na nossa matriz
            if num_drop in matriz_hardware and num_slot in matriz_hardware[num_drop].slots:
                obj_slot = matriz_hardware[num_drop].slots[num_slot]
                
                # Navega na estrutura interna: DIS_CH_IN -> [n] -> VALUE -> Alias
                # Usamos .//instanceElementDesc para buscar os níveis de canais
                for ch_desc in var.findall(".//instanceElementDesc"):
                    ch_name = ch_desc.get("name", "") # Ex: "[0]", "[1]"
                    
                    if ch_name.startswith("[") and ch_name.endswith("]"):
                        try:
                            # Extrai o índice do canal: [0] -> 0
                            indice_canal = int(ch_name.strip("[]"))
                            
                            # Busca o atributo 'Alias' dentro do elemento 'VALUE'
                            value_elem = ch_desc.find(".//instanceElementDesc[@name='VALUE']")
                            if value_elem is not None:
                                attribute_alias = value_elem.find("attribute[@name='Alias']")
                                if attribute_alias is not None:
                                    tag_nome = attribute_alias.get("value")
                                    
                                    # Atribui ao objeto Canal correspondente
                                    # Nota: No seu código Canal(i+1), então indice 0 é Canal(1)
                                    if indice_canal < len(obj_slot.canais):
                                        obj_slot.canais[indice_canal].nome = tag_nome
                                        #print(f"Preenchido: Drop {num_drop} Slot {num_slot} Canal {indice_canal} -> {tag_nome}")
                        except ValueError:
                            continue

def preencher_comentarios_na_matriz(matriz_hardware, lista_variaveis_lidas):
    # 1. Criar um dicionário de busca rápida {nome_da_tag: comentario}
    # Isso evita ter que percorrer a lista inteira para cada canal (O(1) vs O(n))
    mapa_comentarios = {
        var['nome']: var['comentario'] 
        for var in lista_variaveis_lidas 
        if var['nome']
    }

    print("Iniciando preenchimento de comentários...")
    contador = 0

    # 2. Navegar na matriz de hardware
    for drop in matriz_hardware.values():
        for slot in drop.slots.values():
            for canal in slot.canais:
                # Se o canal tiver um nome (tag) atribuído
                if canal.nome:
                    # Busca o comentário no mapa que criamos
                    comentario = mapa_comentarios.get(canal.nome)
                    
                    if comentario:
                        canal.comentario = comentario
                        contador += 1

    print(f"Sucesso: {contador} comentários vinculados aos canais.")


def ler_titulo_modelo(caminho_arquivo_xef: str, lista_variaveis_lidas: List[Dict[str, Any]]) -> str:
 
    
    try:
        tree = ET.parse(caminho_arquivo_xef)
        root = tree.getroot()
        # Busca o partItem que está dentro de PLC
        plc_part = root.find(".//PLC/partItem")
        if plc_part is not None:
            MODELO = plc_part.get("family", "Modelo Desconhecido")
    except Exception as e:
        print(f"Erro ao extrair família do PLC: {e}")
        MODELO = "PLC"

    """
    Lê o atributo 'name' da tag contentHeader no arquivo XEF.
    Se o título for "Project", procura por uma tag terminada em '_DCOM' E do tipo 'WORD'
    na lista de variáveis e a utiliza como título.
    """

    # --- 1. Lógica Original de Leitura do Título no XML ---
    
    titulo_lido = 'Projeto_Invalido'
    
    try:
        tree = ET.parse(caminho_arquivo_xef)
        root = tree.getroot()
        header = root.find('contentHeader')
        
        if header is not None:
            # Pega o nome. Se não houver, usa 'Projeto_Sem_Nome'.
            titulo_lido = header.get('name', 'Projeto_Sem_Nome')
        else:
            titulo_lido = 'Projeto_Sem_Header'
            
    except (FileNotFoundError, ET.ParseError):
        # Mantém 'Projeto_Invalido'
        pass
        
    # --- 2. Lógica de Verificação e Substituição para "_DCOM" e tipo "WORD" ---
    
    if titulo_lido == "Project":
        
        print("\nAlerta: Título original encontrado é 'Project'. Buscando fallback '_DCOM' (Tipo WORD)...")
        
        # Procura a primeira variável que atenda a ambas as condições
        for variavel in lista_variaveis_lidas:
            nome_variavel = variavel.get('nome', '')
            tipo_variavel = variavel.get('tipo', '')
            
            # Condição A: A tag deve terminar com "_DCOM"
            condicao_dcom = nome_variavel and nome_variavel.endswith('_DCOM')
            
            # Condição B: O tipo deve ser "WORD"
            condicao_word = tipo_variavel == 'WORD'
            
            # Verifica se AMBAS as condições são atendidas
            if condicao_dcom and condicao_word:
                print(f"Substituindo 'Project' pela tag: {nome_variavel}")
                return nome_variavel.removesuffix('_DCOM') # Retorna imediatamente o novo título
                
        # 3. Se o loop terminar sem encontrar a tag "_DCOM" tipo "WORD"
        print("Aviso: Nenhuma tag '_DCOM' do tipo 'WORD' foi localizada na lista de variáveis lidas.")
        return titulo_lido
        
    else:
        # Se o título original for válido e não for "Project", retorna o que foi lido
        return titulo_lido, MODELO



#----------------------GERAÇÃO DO ARQUIVO EXCEL----------------------------

def gerar_excel(matriz_hardware, titulo_projeto, modelo_plc):
    data_hoje = datetime.now().strftime("%Y-%m-%d")
    nome_arquivo = f"REMOTE_IO_{titulo_projeto.upper()}_{data_hoje}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de IO"

    # --- CONFIGURAÇÕES DE IMPRESSÃO ---
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0 
    
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Estilos
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True, size=10)
    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_aligned = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

    linha_atual = 1

    for num_drop in sorted(matriz_hardware.keys()):
        obj_drop = matriz_hardware[num_drop]
        for num_slot in sorted(obj_drop.slots.keys()):
            obj_slot = obj_drop.slots[num_slot]

            # --- CABEÇALHO (Linha 1 do Slot) ---
            # Mescla a Coluna 1 (A) e 2 (B) para o texto "VALE"
            ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=2)
            ws.cell(row=linha_atual, column=1, value="VALE").font = header_font
            ws.cell(row=linha_atual, column=3, value=titulo_projeto).font = header_font
            ws.cell(row=linha_atual, column=4, value=f"Modelo\n{modelo_plc}").font = header_font
            ws.cell(row=linha_atual, column=5, value=f"Cartão\n{obj_slot.modelo}").font = header_font
            ws.cell(row=linha_atual, column=6, value=f"Drop\n{num_drop:02d}").font = header_font
            ws.cell(row=linha_atual, column=7, value=f"Slot\n{num_slot:02d}").font
            

            for col in range(1, 8):
                cell = ws.cell(row=linha_atual, column=col)
                cell.alignment = center_aligned
                cell.border = thin_border

            # --- LINHA 2 (Subtítulo e Revisão) ---
            ws.merge_cells(start_row=linha_atual+1, start_column=1, end_row=linha_atual+1, end_column=5)
            ws.cell(row=linha_atual+1, column=1, value="Entradas/Saídas Digitais ou Analógicas").alignment = center_aligned
            ws.merge_cells(start_row=linha_atual+1, start_column=6, end_row=linha_atual+1, end_column=7)
            ws.cell(row=linha_atual+1, column=6, value=f"Revisão: {data_hoje}").alignment = center_aligned
            
            for col in range(1, 8):
                ws.cell(row=linha_atual+1, column=col).border = thin_border

            # --- LINHA 3 (Títulos da Tabela) ---
            ws.cell(row=linha_atual+2, column=1, value="BORNE").font = header_font
            ws.cell(row=linha_atual+2, column=2, value="BIT").font = header_font # Nova Coluna
            ws.cell(row=linha_atual+2, column=3, value="TAG Equipamento").font = header_font
            ws.merge_cells(start_row=linha_atual+2, start_column=4, end_row=linha_atual+2, end_column=7)
            ws.cell(row=linha_atual+2, column=4, value="DESCRIÇÃO / COMENTÁRIO").font = header_font
            
            for col in range(1, 8):
                cell = ws.cell(row=linha_atual+2, column=col)
                cell.alignment = center_aligned
                cell.border = thin_border
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            # --- CANAIS (Preenchimento) ---
            for i in range(32):
                r_idx = linha_atual + 3 + i
                ws.row_dimensions[r_idx].height = 14.5 # Ajuste para caber no A4 Paisagem
                
                # Borne (Coluna 1)
                ws.cell(row=r_idx, column=1, value=i+1).alignment = center_aligned
                
                # Bit (Coluna 2: Borne - 1)
                ws.cell(row=r_idx, column=2, value=i).alignment = center_aligned
                
                tag = "-"
                coment = "-"
                if i < len(obj_slot.canais):
                    tag = obj_slot.canais[i].nome or "-"
                    coment = obj_slot.canais[i].comentario or "-"

                # Tag (Coluna 3)
                ws.cell(row=r_idx, column=3, value=tag).alignment = center_aligned
                
                # Comentário (Colunas 4 a 7 mescladas)
                ws.merge_cells(start_row=r_idx, start_column=4, end_row=r_idx, end_column=7)
                ws.cell(row=r_idx, column=4, value=coment).alignment = left_aligned

                for col in range(1, 8):
                    ws.cell(row=r_idx, column=col).border = thin_border

            # --- FINALIZAÇÃO DO SLOT ---
            linha_atual += 35 
            ws.row_breaks.append(Break(id=linha_atual-1))

    # Ajuste final de colunas (A=Borne, B=Bit, C=Tag, D...G=Comentário)
    larguras = [10, 10, 30, 21, 21, 21, 21] 
    for i, w in enumerate(larguras):
        ws.column_dimensions[chr(65+i)].width = w

    wb.save(nome_arquivo)
    print(f"Arquivo único gerado: {nome_arquivo}")
#----------------------MAIN----------------------------
if __name__ == "__main__":

    # --- DEFINIÇÃO UNIVERSAL DO CAMINHO BASE ---
    # Essa lógica funciona tanto para o script .py quanto para o executável .exe (frozen)
    if getattr(sys, 'frozen', False):
        # Se estiver rodando como executável (PyInstaller), usa o caminho do binário.
        diretorio_script = os.path.dirname(sys.executable)
    else:
        # Se estiver rodando como script Python (.py), usa o caminho do arquivo de script.
        # É fundamental usar o try-except ou um método robusto para evitar erros ao ser chamado de outro diretório.
        try:
            diretorio_script = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            # Fallback caso __file__ não esteja definido (raro, mas seguro)
            diretorio_script = os.path.getcwd() 

    # --- Configuração de Caminhos ---
    caminho_unitpro = os.path.join(diretorio_script, ARQUIVO_UNITPRO)

    # 1. Leitura e Catalogação das Variáveis
    lista_variaveis_lidas = ler_variaveis_unitpro(caminho_unitpro)

    # 2. Gerar a estrutura a partir do hardware do PLC
    matriz_hardware = gerar_matriz_plc(caminho_unitpro)
 
    # 3. Preencher os nomes dos canais com base nas variáveis do arquivo
    preencher_canais_da_matriz(caminho_unitpro, matriz_hardware)

    # 4. Preencher os COMENTÁRIOS nos canais
    # Cruza os dados da matriz com a lista_variaveis_lidas
    preencher_comentarios_na_matriz(matriz_hardware, lista_variaveis_lidas)

    # Supondo que você extraiu essas informações do XML ou entrada do usuário:
    titulo_projeto, modelo_plc = ler_titulo_modelo(caminho_unitpro,lista_variaveis_lidas) 
    

    # 6. Geração do arquivo com o nome dinâmico: REMOTE_IO_[UC1000CC21]_2025-12-31.xlsx
    gerar_excel(matriz_hardware, titulo_projeto, modelo_plc)
    print("Processamento concluído.") 